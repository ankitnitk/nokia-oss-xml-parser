#!/usr/bin/env python3
"""
OSS XML to XLSX Converter  — Version 2  (Parallel Sheet Writing)
Converts Nokia OSS RAML XML dump files to Excel format.

Key improvement over V1: all sheets are written simultaneously in separate
processes, then merged — write time ~= slowest single sheet instead of sum.

Usage (interactive GUI):   python oss_xml_to_xlsx_v2.py
Usage (command-line):      python oss_xml_to_xlsx_v2.py file1.xml.gz file2.gz -o out.xlsx
                           python oss_xml_to_xlsx_v2.py file1.zip --classes BTS,BCF -o out.xlsx
"""

import gzip
import io
import os
import re
import sys
import zipfile
import argparse
import threading
from collections import defaultdict, OrderedDict
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from datetime import datetime
from multiprocessing import freeze_support

from lxml import etree
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

try:
    import win32com.client
    XLSB_SUPPORTED = True
except ImportError:
    XLSB_SUPPORTED = False

AUTHOR = 'Ankit Jain'

# Match by local name only — handles raml20.xsd, raml21.xsd, any future versions
def _local(tag):
    return tag.split('}', 1)[-1] if '}' in tag else tag

TAG_MO   = 'managedObject'
TAG_P    = 'p'
TAG_LIST = 'list'
TAG_ITEM = 'item'
TAG_LOG  = 'log'

# Styles — defined as dicts so they can be passed to worker processes
_H_FONT  = dict(bold=True, color='FFFFFFFF')
_H_FILL  = dict(patternType='solid', fgColor='FF4472C4')
_H_ALIGN = dict(vertical='center')
_H_BORDER_SIDE = dict(border_style='thin')
_LINK_FONT = dict(color='FF0563C1', underline='single')

_print_lock = threading.Lock()
def tprint(*a, **k):
    with _print_lock:
        print(*a, **k)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fmt_elapsed(s):
    if s < 60: return f'{s:.1f}s'
    m, sec = divmod(int(s), 60)
    return f'{m}m {sec:02d}s'

def ts():
    return datetime.now().strftime('%H:%M:%S')


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------

def iter_xml_streams(filepath):
    name = os.path.basename(filepath)
    low  = filepath.lower()
    if low.endswith('.zip'):
        with zipfile.ZipFile(filepath, 'r') as zf:
            entries = [e for e in zf.namelist()
                       if e.lower().endswith('.xml') and not e.startswith('__MACOSX')]
            if not entries:
                raise ValueError(f'No .xml files found inside {name}')
            for entry in entries:
                with zf.open(entry) as f:
                    yield f.read(), f'{name}/{os.path.basename(entry)}'
    elif low.endswith('.gz'):
        with gzip.open(filepath, 'rb') as f:
            yield f.read(), name
    else:
        with open(filepath, 'rb') as f:
            yield f.read(), name


# ---------------------------------------------------------------------------
# Quick scan
# ---------------------------------------------------------------------------

_CLASS_RE = re.compile(rb'managedObject class="([^"]+)"')

def quick_scan_classes(filepath):
    counts = defaultdict(int)
    low = filepath.lower()
    if low.endswith('.zip'):
        with zipfile.ZipFile(filepath, 'r') as zf:
            for entry in [e for e in zf.namelist() if e.lower().endswith('.xml')]:
                with zf.open(entry) as f:
                    for m in _CLASS_RE.finditer(f.read()):
                        counts[m.group(1).decode()] += 1
    elif low.endswith('.gz'):
        with gzip.open(filepath, 'rb') as f:
            for chunk in iter(lambda: f.read(1 << 20), b''):
                for m in _CLASS_RE.finditer(chunk):
                    counts[m.group(1).decode()] += 1
    else:
        with open(filepath, 'rb') as f:
            for chunk in iter(lambda: f.read(1 << 20), b''):
                for m in _CLASS_RE.finditer(chunk):
                    counts[m.group(1).decode()] += 1
    return counts

def scan_all_files(filepaths):
    merged = defaultdict(int)
    with ThreadPoolExecutor(max_workers=min(len(filepaths), 4)) as ex:
        for counts in ex.map(quick_scan_classes, filepaths):
            for cls, cnt in counts.items():
                merged[cls] += cnt
    return merged


# ---------------------------------------------------------------------------
# XML Parsing
# ---------------------------------------------------------------------------

def try_numeric(text):
    if text is None: return None
    s = text.strip()
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return s or None

def parse_dist_name(dist_name):
    h = OrderedDict()
    for part in dist_name.split('/'):
        if '-' not in part: continue
        idx = part.index('-')
        cls, oid = part[:idx], part[idx + 1:]
        if cls == 'PLMN': continue
        try:
            f = float(oid)
            h[cls] = int(f) if f == int(f) else f
        except ValueError:
            h[cls] = oid
    return h

def parse_managed_object(elem, filename):
    mo_class  = elem.get('class', '')
    dist_name = elem.get('distName', '')
    obj_id    = elem.get('id', '')
    version   = elem.get('version', '')
    operation = elem.get('operation', '')  # present in plan/change XMLs
    hierarchy = parse_dist_name(dist_name)
    record = OrderedDict()
    if operation:
        record['operation']  = operation
    record['id']         = try_numeric(obj_id)
    record['File_Name']  = filename
    record['Dist_Name']  = dist_name
    record['SW_Version'] = version
    for child in elem:
        ltag = _local(child.tag)
        if ltag == TAG_P:
            record[child.get('name')] = try_numeric(child.text)
        elif ltag == TAG_LIST:
            list_name = child.get('name')
            children  = list(child)
            if not children:
                pass
            elif _local(children[0].tag) == TAG_ITEM:
                record[list_name] = 'List'
                item_fields = OrderedDict()
                for item in children:
                    for p in item:
                        if _local(p.tag) == TAG_P:
                            item_fields.setdefault(p.get('name'), []).append(str(p.text or ''))
                for fname, vals in item_fields.items():
                    record[f'Item-{list_name}-{fname}'] = ';'.join(vals)
            else:
                vals = [str(p.text or '') for p in children if _local(p.tag) == TAG_P]
                record[list_name] = 'List;' + ';'.join(vals)
    return mo_class, hierarchy, record

def parse_xml_bytes(data, display_name, filter_classes=None):
    classes_data = defaultdict(list)
    file_info = {'filename': display_name, 'dateTime': ''}
    count = 0
    t0 = datetime.now()
    tprint(f'  [{ts()}] Parsing {display_name} ...', end='', flush=True)
    stream = io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data
    for event, elem in etree.iterparse(stream, events=('end',),
                                       load_dtd=False, no_network=True,
                                       recover=True, huge_tree=True):
        ltag = _local(elem.tag)
        if ltag == TAG_LOG and not file_info['dateTime']:
            file_info['dateTime'] = elem.get('dateTime', '')
            elem.clear()
        elif ltag == TAG_MO:
            mo_class = elem.get('class', '')
            if filter_classes is None or mo_class in filter_classes:
                _, hierarchy, record = parse_managed_object(elem, display_name)
                classes_data[mo_class].append((hierarchy, record))
                count += 1
            elem.clear()
    elapsed = (datetime.now() - t0).total_seconds()
    tprint(f' done — {count:,} objects in {fmt_elapsed(elapsed)}')
    return classes_data, file_info

def parse_input_file(filepath, filter_classes=None):
    results = []
    for xml_bytes, display_name in iter_xml_streams(filepath):
        results.append(parse_xml_bytes(xml_bytes, display_name, filter_classes))
    return results


# ---------------------------------------------------------------------------
# Column ordering + record flattening
# ---------------------------------------------------------------------------

def build_column_order(records):
    META = ('operation', 'id', 'File_Name', 'Dist_Name', 'SW_Version')
    hier_seen, param_seen = OrderedDict(), OrderedDict()
    has_op = False
    for hierarchy, record in records:
        for k in hierarchy: hier_seen[k] = True
        for k in record:
            if k == 'operation': has_op = True
            if k not in META: param_seen[k] = True
    meta = list(META) if has_op else [m for m in META if m != 'operation']
    return list(hier_seen.keys()), meta, list(param_seen.keys())

def flatten_records(records, all_cols, hier_set):
    """Convert list of (hierarchy, record) to list of plain lists for fast pickling."""
    col_info = [(col in hier_set, col) for col in all_cols]
    return [
        [hierarchy.get(col) if is_h else record.get(col)
         for is_h, col in col_info]
        for hierarchy, record in records
    ]


# ---------------------------------------------------------------------------
# Sheet writing worker  (runs in separate process)
# ---------------------------------------------------------------------------

def _make_header_cell(ws, value):
    cell = WriteOnlyCell(ws, value=value)
    cell.font      = Font(**_H_FONT)
    cell.fill      = PatternFill(**_H_FILL)
    cell.alignment = Alignment(**_H_ALIGN)
    side = Side(**_H_BORDER_SIDE)
    cell.border = Border(left=side, right=side, top=side, bottom=side)
    return cell

def _make_link_cell(ws, value, location):
    cell = WriteOnlyCell(ws, value=value)
    cell.font      = Font(**_LINK_FONT)
    cell.hyperlink = Hyperlink(ref='', location=location)
    return cell

def write_sheet_worker(args):
    """Runs in a subprocess. Writes one MO class sheet to xlsx bytes."""
    cls, flat_records, all_cols, n_hier = args
    t_start = datetime.now()

    buf    = io.BytesIO()
    wb     = Workbook(write_only=True)
    ws     = wb.create_sheet(cls)
    n      = n_hier
    n_cols = len(all_cols)
    has_op = 'operation' in all_cols

    # Freeze after Dist_Name — shifts by 1 if operation column present
    dist_freeze = n + (1 if has_op else 0) + 3 + 1   # 1-based column after Dist_Name
    ws.freeze_panes = f'{get_column_letter(dist_freeze)}3'
    ws.auto_filter.ref = f'A2:{get_column_letter(n_cols)}2'

    # Column widths
    for i in range(n):
        ws.column_dimensions[get_column_letter(i + 1)].width = 12  # hierarchy
    ci = n + 1
    if has_op:
        ws.column_dimensions[get_column_letter(ci)].width = 10; ci += 1  # operation
    ws.column_dimensions[get_column_letter(ci)].width = 12;     ci += 1  # id
    ws.column_dimensions[get_column_letter(ci)].width = 22;     ci += 1  # File_Name
    ws.column_dimensions[get_column_letter(ci)].width = 45;     ci += 1  # Dist_Name
    ws.column_dimensions[get_column_letter(ci)].width = 10;     ci += 1  # SW_Version
    for i in range(ci, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Row 0: "Info" link back to Info sheet
    info_row = [_make_link_cell(ws, 'Info', "'Info'!A1")]
    info_row += [None] * (n_cols - 1)
    ws.append(info_row)

    # Row 1: headers
    ws.append([_make_header_cell(ws, col) for col in all_cols])

    # Rows 2+: data
    for row in flat_records:
        ws.append(row)

    wb.save(buf)
    elapsed = (datetime.now() - t_start).total_seconds()
    return cls, buf.getvalue(), elapsed


# ---------------------------------------------------------------------------
# xlsx merge  (combines single-sheet workbooks into one)
# ---------------------------------------------------------------------------

WS_CT  = 'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'
WB_CT  = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml'
STY_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml'
WS_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet'
STY_REL= 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles'
DOC_REL= 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument'

def _content_types_xml(n_sheets):
    overrides = ''.join(
        f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" ContentType="{WS_CT}"/>'
        for i in range(n_sheets)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        f'<Override PartName="/xl/workbook.xml" ContentType="{WB_CT}"/>'
        f'<Override PartName="/xl/styles.xml" ContentType="{STY_CT}"/>'
        + overrides +
        '</Types>'
    ).encode()

def _root_rels_xml():
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Type="{DOC_REL}" Target="xl/workbook.xml"/>'
        '</Relationships>'
    ).encode()

def _workbook_xml(sheet_names):
    sheets = ''.join(
        f'<sheet name="{name}" sheetId="{i+1}" r:id="rId{i+1}"/>'
        for i, name in enumerate(sheet_names)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<bookViews><workbookView activeTab="0"/></bookViews>'
        f'<sheets>{sheets}</sheets>'
        '</workbook>'
    ).encode()

def _workbook_rels_xml(n_sheets):
    ws_rels = ''.join(
        f'<Relationship Id="rId{i+1}" Type="{WS_REL}" Target="worksheets/sheet{i+1}.xml"/>'
        for i in range(n_sheets)
    )
    sty_rel = f'<Relationship Id="rId{n_sheets+1}" Type="{STY_REL}" Target="styles.xml"/>'
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + ws_rels + sty_rel +
        '</Relationships>'
    ).encode()

def _docprops_core_xml():
    now = datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"'
        ' xmlns:dc="http://purl.org/dc/elements/1.1/"'
        ' xmlns:dcterms="http://purl.org/dc/terms/"'
        ' xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        '<dc:title>OSS XML Dump</dc:title>'
        f'<dc:creator>{AUTHOR}</dc:creator>'
        f'<cp:lastModifiedBy>{AUTHOR}</cp:lastModifiedBy>'
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>'
        '</cp:coreProperties>'
    ).encode()

def merge_xlsx(sheet_order, sheet_xlsx_map, output_path):
    """
    sheet_order: list of sheet names in order
    sheet_xlsx_map: {name: xlsx_bytes}
    Merges all into one workbook at output_path.
    """
    # Extract worksheet XML, rels, and styles from each sub-workbook
    worksheets = []  # [(name, ws_xml_bytes, ws_rels_bytes_or_None)]
    styles_xml = None

    for name in sheet_order:
        with zipfile.ZipFile(io.BytesIO(sheet_xlsx_map[name]), 'r') as zf:
            names = zf.namelist()
            ws_xml  = zf.read('xl/worksheets/sheet1.xml')
            ws_rels = zf.read('xl/worksheets/_rels/sheet1.xml.rels') \
                      if 'xl/worksheets/_rels/sheet1.xml.rels' in names else None
            worksheets.append((name, ws_xml, ws_rels))
            # Take styles from a data sheet — Info sheet uses different styles
            # and its indices don't match data sheet workers
            if styles_xml is None and 'xl/styles.xml' in names and name != 'Info':
                styles_xml = zf.read('xl/styles.xml')

    n = len(worksheets)

    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as out:
        out.writestr('[Content_Types].xml',        _content_types_xml(n))
        out.writestr('_rels/.rels',                _root_rels_xml())
        out.writestr('xl/workbook.xml',            _workbook_xml(sheet_order))
        out.writestr('xl/_rels/workbook.xml.rels', _workbook_rels_xml(n))
        out.writestr('xl/styles.xml',              styles_xml)
        out.writestr('docProps/core.xml',          _docprops_core_xml())
        for i, (name, ws_xml, ws_rels) in enumerate(worksheets):
            out.writestr(f'xl/worksheets/sheet{i+1}.xml', ws_xml)
            if ws_rels:
                out.writestr(f'xl/worksheets/_rels/sheet{i+1}.xml.rels', ws_rels)


# ---------------------------------------------------------------------------
# Info sheet writer  (runs in main process)
# ---------------------------------------------------------------------------

def write_info_sheet_bytes(files_info, class_names):
    """Write the Info sheet to xlsx bytes using openpyxl write_only."""
    buf = io.BytesIO()
    wb  = Workbook(write_only=True)
    ws  = wb.create_sheet('Info')

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 38

    bold = Font(bold=True)
    link = Font(**_LINK_FONT)

    def lbl(v):
        c = WriteOnlyCell(ws, value=v); c.font = bold; return c
    def txt(v):
        return WriteOnlyCell(ws, value=v)
    def hyperlink_cell(v, location):
        c = WriteOnlyCell(ws, value=v)
        c.font = link
        c.hyperlink = Hyperlink(ref='', location=location)
        return c

    ws.append([None])
    ws.append([None, lbl('Created with OSS XML Converter  —  V2')])
    ws.append([None, txt(f'Created by: {AUTHOR}')])
    ws.append([None])
    ws.append([None, None, lbl('File name'), lbl('Dump Extracted Time')])
    for fi in files_info:
        ws.append([None, txt('Used Netact export:'), txt(fi['filename']), txt(fi['dateTime'])])
    ws.append([None])
    ws.append([None, lbl('Contents')])
    for cls in class_names:
        ws.append([None, hyperlink_cell(cls, f"'{cls}'!A1")])

    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# XLSB conversion
# ---------------------------------------------------------------------------

def convert_to_xlsb(xlsx_path, xlsb_path):
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.SaveAs(os.path.abspath(xlsb_path), FileFormat=50)
        wb.Close(False)
    finally:
        excel.Quit()


# ---------------------------------------------------------------------------
# Core conversion
# ---------------------------------------------------------------------------

def run_conversion(input_paths, output_path, filter_classes=None):
    t_total = datetime.now()

    want_xlsb = output_path.lower().endswith('.xlsb')
    if want_xlsb and not XLSB_SUPPORTED:
        print('WARNING: .xlsb requires pywin32 (pip install pywin32). Falling back to .xlsx.')
        output_path = output_path[:-5] + '.xlsx'
        want_xlsb   = False

    xlsx_path = output_path if not want_xlsb else output_path[:-5] + '_tmp.xlsx'

    print(f'[{ts()}] Output: {output_path}')
    if filter_classes:
        print(f'[{ts()}] Filtering to classes: {", ".join(sorted(filter_classes))}')
    print(f'[{ts()}] Parsing {len(input_paths)} file(s) in parallel...')

    # --- Parse ---
    all_cd, all_fi = [], []
    t_parse = datetime.now()
    with ThreadPoolExecutor(max_workers=min(len(input_paths), 4)) as ex:
        futs = {ex.submit(parse_input_file, p, filter_classes): p for p in input_paths}
        for fut in as_completed(futs):
            for cd, fi in fut.result():
                all_cd.append(cd)
                all_fi.append(fi)
    t_parse = (datetime.now() - t_parse).total_seconds()

    merged = defaultdict(list)
    for cd in all_cd:
        for cls, recs in cd.items():
            merged[cls].extend(recs)

    # Re-order file info to match input order
    fi_by_name = {fi['filename']: fi for fi in all_fi}
    ordered_fi = []
    for path in input_paths:
        name = os.path.basename(path)
        if name in fi_by_name and fi_by_name[name] not in ordered_fi:
            ordered_fi.append(fi_by_name[name])
    for fi in all_fi:
        if fi not in ordered_fi:
            ordered_fi.append(fi)

    class_names   = sorted(merged.keys())
    total_records = sum(len(v) for v in merged.values())
    print(f'[{ts()}] Parsing done in {fmt_elapsed(t_parse)} — '
          f'{total_records:,} records, {len(class_names)} classes: {", ".join(class_names)}')

    # --- Prepare worker args ---
    worker_args = []
    for cls in class_names:
        recs = merged[cls]
        hier_cols, meta_cols, param_cols = build_column_order(recs)
        all_cols  = hier_cols + meta_cols + param_cols
        hier_set  = set(hier_cols)
        flat      = flatten_records(recs, all_cols, hier_set)
        worker_args.append((cls, flat, all_cols, len(hier_cols)))

    # --- Parallel write ---
    print(f'[{ts()}] Writing {len(class_names)} sheets in parallel...')
    t_write = datetime.now()

    sheet_xlsx = {}   # cls -> xlsx_bytes

    # Info sheet in main process (fast)
    print(f'  [{ts()}] Sheet "Info" ...', end='', flush=True)
    t0 = datetime.now()
    sheet_xlsx['Info'] = write_info_sheet_bytes(ordered_fi, class_names)
    print(f' done ({fmt_elapsed((datetime.now()-t0).total_seconds())})')

    # Data sheets in parallel processes
    n_workers = min(len(worker_args), os.cpu_count() or 4)

    if not worker_args:
        print(f'  [{ts()}] No data sheets to write.')
    else:
        print(f'  [{ts()}] Launching {n_workers} parallel workers...')

    with ProcessPoolExecutor(max_workers=max(n_workers, 1)) as ex:
        futs = {ex.submit(write_sheet_worker, args): args[0] for args in worker_args}
        for fut in as_completed(futs):
            cls = futs[fut]
            try:
                cls_result, xlsx_bytes, elapsed = fut.result()
                sheet_xlsx[cls_result] = xlsx_bytes
                print(f'  [{ts()}] Sheet "{cls_result}" done ({fmt_elapsed(elapsed)})')
            except Exception as e:
                print(f'  [{ts()}] Sheet "{cls}" FAILED: {e}')

    t_write = (datetime.now() - t_write).total_seconds()
    print(f'[{ts()}] All sheets written in {fmt_elapsed(t_write)}')

    # --- Merge ---
    print(f'[{ts()}] Merging into {os.path.basename(xlsx_path)} ...', end='', flush=True)
    t0 = datetime.now()
    sheet_order = ['Info'] + class_names
    merge_xlsx(sheet_order, sheet_xlsx, xlsx_path)
    t_merge = (datetime.now() - t0).total_seconds()
    print(f' done ({fmt_elapsed(t_merge)})')

    if want_xlsb:
        t0 = datetime.now()
        print(f'[{ts()}] Converting to .xlsb via Excel...')
        convert_to_xlsb(xlsx_path, output_path)
        os.remove(xlsx_path)
        print(f'[{ts()}] Conversion done in {fmt_elapsed((datetime.now()-t0).total_seconds())}')

    t_total = (datetime.now() - t_total).total_seconds()
    size_mb = os.path.getsize(output_path) / 1024 / 1024
    print(f'\n[{ts()}] Done!  {output_path}  ({size_mb:.1f} MB)')
    print(f'         Parse: {fmt_elapsed(t_parse)}  |  '
          f'Write: {fmt_elapsed(t_write)}  |  '
          f'Merge: {fmt_elapsed(t_merge)}  |  '
          f'Total: {fmt_elapsed(t_total)}')
    return output_path


# ---------------------------------------------------------------------------
# Class selection dialog
# ---------------------------------------------------------------------------

def ask_class_selection(class_counts):
    import tkinter as tk
    selected = {}
    result   = [None]

    dlg = tk.Toplevel()
    dlg.title('Select MO Classes')
    dlg.resizable(False, False)
    dlg.grab_set()

    total = sum(class_counts.values())
    tk.Label(dlg, text=f'{len(class_counts)} MO classes  |  {total:,} total objects',
             font=('Arial', 10, 'bold'), pady=6).pack()

    container = tk.Frame(dlg)
    container.pack(padx=12, pady=(0, 4), fill='both')
    canvas    = tk.Canvas(container, width=340, height=280, bd=1,
                          relief='sunken', highlightthickness=0)
    scrollbar = tk.Scrollbar(container, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side='right', fill='y')
    canvas.pack(side='left', fill='both')
    inner = tk.Frame(canvas)
    cw    = canvas.create_window((0, 0), window=inner, anchor='nw')

    def on_resize(event):
        canvas.configure(scrollregion=canvas.bbox('all'))
        canvas.itemconfig(cw, width=event.width)
    inner.bind('<Configure>', on_resize)
    canvas.bind('<MouseWheel>', lambda e: canvas.yview_scroll(int(-1 * e.delta / 120), 'units'))

    for i, cls in enumerate(sorted(class_counts)):
        var = tk.BooleanVar(value=True)
        selected[cls] = var
        row, col = divmod(i, 2)
        tk.Checkbutton(inner, text=f'{cls}  ({class_counts[cls]:,})',
                       variable=var, anchor='w',
                       font=('Consolas', 9), width=18).grid(
            row=row, column=col, sticky='w', padx=6, pady=1)

    btn_frame = tk.Frame(dlg)
    btn_frame.pack(pady=6)

    def ok():
        result[0] = {c for c, v in selected.items() if v.get()}
        dlg.destroy()

    tk.Button(btn_frame, text='All',    width=8,
              command=lambda: [v.set(True)  for v in selected.values()]).grid(row=0, column=0, padx=3)
    tk.Button(btn_frame, text='None',   width=8,
              command=lambda: [v.set(False) for v in selected.values()]).grid(row=0, column=1, padx=3)
    tk.Button(btn_frame, text='OK',     width=8,  command=ok,
              bg='#4472C4', fg='white').grid(row=0, column=2, padx=3)
    tk.Button(btn_frame, text='Cancel', width=8,
              command=dlg.destroy).grid(row=0, column=3, padx=3)

    dlg.wait_window()
    return result[0]


# ---------------------------------------------------------------------------
# Interactive GUI
# ---------------------------------------------------------------------------

def run_interactive():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()

    input_paths = list(filedialog.askopenfilenames(
        title='Select OSS XML dump file(s)',
        filetypes=[('All supported', '*.xml *.gz *.zip'),
                   ('XML files', '*.xml'), ('GZ files', '*.gz'),
                   ('ZIP files', '*.zip'), ('All files', '*.*')],
    ))
    if not input_paths:
        print('No files selected.'); return

    print(f'[{ts()}] Scanning for MO classes...')
    t0 = datetime.now()
    class_counts = scan_all_files(input_paths)
    print(f'[{ts()}] Scan done in {fmt_elapsed((datetime.now()-t0).total_seconds())} '
          f'— classes: {", ".join(sorted(class_counts))}')

    filter_classes = ask_class_selection(class_counts)
    if not filter_classes:
        print('No classes selected.'); return

    out_types = ([('Excel Binary Workbook', '*.xlsb')] if XLSB_SUPPORTED else []) + \
                [('Excel Workbook', '*.xlsx')]
    base    = os.path.basename(input_paths[0]).split('.')[0]
    defname = f'{base}_{datetime.now().strftime("%Y%m%d_%H%M%S")}_dump.xlsx'
    output_path = filedialog.asksaveasfilename(
        title='Save output as', initialdir=os.path.dirname(input_paths[0]),
        initialfile=defname, defaultextension='.xlsx', filetypes=out_types,
    )
    if not output_path:
        print('No output file selected.'); return

    root.destroy()
    run_conversion(input_paths, output_path, filter_classes or None)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    os.system("title Ankit's XML Parser  [V2 - Parallel]")

    if len(sys.argv) == 1:
        run_interactive()
        input('\nPress Enter to exit...')
        return

    parser = argparse.ArgumentParser(description='OSS XML → Excel Converter V2 (Parallel)')
    parser.add_argument('inputs',    nargs='+', metavar='FILE')
    parser.add_argument('-o', '--output',  default='')
    parser.add_argument('--classes', default='')
    args = parser.parse_args()

    for path in args.inputs:
        if not os.path.isfile(path):
            print(f'ERROR: File not found: {path}', file=sys.stderr)
            input('\nPress Enter to exit...')
            sys.exit(1)

    if not args.output:
        base = os.path.basename(args.inputs[0]).split('.')[0]
        args.output = f'{base}_{datetime.now().strftime("%Y%m%d_%H%M%S")}_dump.xlsx'

    filter_classes = {c.strip() for c in args.classes.split(',') if c.strip()} or None
    run_conversion(args.inputs, args.output, filter_classes)
    input('\nPress Enter to exit...')


if __name__ == '__main__':
    freeze_support()   # required for PyInstaller + multiprocessing on Windows
    main()
