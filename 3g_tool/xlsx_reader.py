"""
xlsx_reader.py
Reads Nokia-format .xlsx and .xlsb parameter dump files using python-calamine
(fast Rust-based reader). Falls back to openpyxl for .xlsx if calamine is not
installed.

Presents the same interface as xlsb_reader.read_xlsb().
Sheet structure: Row 0 = ignored, Row 1 = headers (trimmed), Row 2+ = data.
Dist_Name is the unique key per sheet (first occurrence kept).
"""


def _clean(val):
    """Normalise a cell value to str, converting whole-number floats cleanly."""
    if val is None:
        return ''
    if isinstance(val, float):
        if val != val:        # NaN
            return ''
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def _records_from_rows(rows_iter, header_row=1):
    """
    Given an iterable of rows (each row = list of values), parse into records.

    header_row : 0-based index of the row that contains column headers.
                 Rows before it are ignored; rows after it are data.
                 Default 1 = Excel row 2 (standard Nokia format).
                 Pass 0 for files whose headers start on Excel row 1.

    Returns list of dicts. Rows without a Dist_Name are kept (Dist_Name may be
    synthesised later by Network._fill_dist_names).
    """
    headers = None
    records = []

    for row_idx, row in enumerate(rows_iter):
        if row_idx < header_row:
            continue                          # rows above the header — ignore
        if row_idx == header_row:
            headers = [_clean(v) for v in row]
            continue
        if headers is None:
            continue

        rec = {}
        for col_idx, hdr in enumerate(headers):
            if not hdr:
                continue
            val = _clean(row[col_idx]) if col_idx < len(row) else ''
            rec[hdr] = val

        records.append(rec)

    return records


# ---------------------------------------------------------------------------
# calamine-based reader (preferred - fast Rust engine, supports xlsx + xlsb)
# ---------------------------------------------------------------------------

def _read_with_calamine(path, sheet_names, progress_fn, header_row):
    from python_calamine import CalamineWorkbook

    def log(msg):
        if progress_fn:
            progress_fn(msg)

    log('Opening workbook with calamine...')
    with open(path, 'rb') as f:
        wb = CalamineWorkbook.from_filelike(f)

    all_sheets = wb.sheet_names
    log(f'  Sheets found: {all_sheets}')

    result = {}
    for name in all_sheets:
        if sheet_names and name not in sheet_names:
            continue
        log(f'  Reading sheet: {name}...')
        rows = wb.get_sheet_by_name(name).to_python()
        records = _records_from_rows(iter(rows), header_row=header_row)
        result[name] = records
        log(f'    {len(records):,} records')

    return result


# ---------------------------------------------------------------------------
# openpyxl fallback (xlsx only)
# ---------------------------------------------------------------------------

def _read_with_openpyxl(path, sheet_names, progress_fn, header_row):
    import openpyxl

    def log(msg):
        if progress_fn:
            progress_fn(msg)

    log('Opening workbook with openpyxl (calamine not installed)...')
    with open(path, 'rb') as f:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    all_sheets = wb.sheetnames
    log(f'  Sheets found: {all_sheets}')

    result = {}
    for name in all_sheets:
        if sheet_names and name not in sheet_names:
            continue
        log(f'  Reading sheet: {name}...')

        def row_iter(ws):
            for row in ws.iter_rows(values_only=True):
                yield list(row)

        ws = wb[name]
        records = _records_from_rows(row_iter(ws), header_row=header_row)
        result[name] = records
        log(f'    {len(records):,} records')

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def read_xlsx(path, sheet_names=None, progress_fn=None, header_row=1):
    """
    Read an .xlsx or .xlsb file using calamine (preferred) or openpyxl fallback.
    Returns dict of {sheet_name: [records]}.

    header_row : 0-based row index of the header row (0 = Excel row 1, 1 = Excel row 2).
    """
    try:
        import python_calamine  # noqa: F401
        return _read_with_calamine(path, sheet_names, progress_fn, header_row)
    except ImportError:
        import os
        if os.path.splitext(path)[1].lower() == '.xlsb':
            raise RuntimeError(
                'python-calamine is required to read .xlsb files.\n'
                'Run: pip install python-calamine'
            )
        return _read_with_openpyxl(path, sheet_names, progress_fn, header_row)
